from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


HEADER_MAP = {
    "Ürün Grup ID": "productGroupId",
    "Varyant ID": "variantId",
    "İsim": "name",
    "Açıklama": "description",
    "Satış Fiyatı": "salePrice",
    "İndirimli Fiyatı": "discountedPrice",
    "Alış Fiyatı": "purchasePrice",
    "Barkod Listesi": "barcodeList",
    "SKU": "sku",
    "Silindi mi?": "isDeleted",
    "Marka": "brand",
    "Kategoriler": "categories",
    "Etiketler": "tags",
    "Resim URL": "imageUrl",
    "Metadata Başlık": "metadataTitle",
    "Metadata Açıklama": "metadataDescription",
    "Slug": "slug",
    "Stok:Ana Depo": "stock",
    "Tip": "type",
    "Varyant Tip 1": "variantType1",
    "Varyant Değer 1": "variantValue1",
    "Varyant Tip 2": "variantType2",
    "Varyant Değer 2": "variantValue2",
    "Desi": "desi",
    "HS Kod": "hsCode",
    "Birim Ürün Miktarı": "unitProductQuantity",
    "Ürün Birimi": "productUnit",
    "Satılan Ürün Miktarı": "soldProductQuantity",
    "Satılan Ürün Birimi": "soldProductUnit",
    "Google ürün Kategorisi": "googleProductCategory",
    "Tedarikçi": "supplier",
    "Stoğu Tükenince Satmaya Devam Et": "continueSellingWhenOutOfStock",
    "Satış Kanalı:vitacure": "salesChannel",
    "Sepet Başına Minimum Alma Adeti:vitacure": "minBasketQuantity",
    "Sepet Başına Maksimum Alma Adeti:vitacure": "maxBasketQuantity",
    "Varyant Aktiflik": "variantIsActive",
    "Oluşturulma Tarihi": "createdAt",
}

BOOLEAN_KEYS = {"isDeleted", "variantIsActive"}
IMAGE_SUFFIXES = (".webp", ".jpg", ".jpeg", ".png", ".gif", ".avif")


def repair_text(text: str | None) -> str | None:
    if text is None:
        return None
    if not any(token in text for token in ("Ã", "Ä", "Å", "â")):
        return text
    try:
        return text.encode("latin-1").decode("utf-8")
    except UnicodeError:
        return text


def resolve_primary_image_url(text: str | None) -> str | None:
    if not text:
        return None

    parts = [part.strip() for part in text.split(";") if part and part.strip()]
    for part in parts:
        lowered = part.lower()
        if lowered.startswith("http") and lowered.endswith(IMAGE_SUFFIXES):
            return part

    return None


def normalize_value(value, key):
    if value is None:
        return None
    if isinstance(value, bool):
        return value if key in BOOLEAN_KEYS else ("true" if value else "false")
    if isinstance(value, (int, float)):
        return str(value).rstrip("0").rstrip(".") if isinstance(value, float) else str(value)
    text = repair_text(str(value).strip())
    if key == "imageUrl":
        return resolve_primary_image_url(text)
    return text or None


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: export_ikas_json.py <input.xlsx> <output.json>")
        return 1

    input_path = Path(sys.argv[1]).expanduser().resolve()
    output_path = Path(sys.argv[2]).expanduser().resolve()
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    rows = worksheet.iter_rows(values_only=True)
    headers = [repair_text(str(cell).strip()) if cell is not None else "" for cell in next(rows)]

    data = []
    for row in rows:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        item = {}
        for index, header in enumerate(headers):
            key = HEADER_MAP.get(header)
            if not key:
                continue
            item[key] = normalize_value(row[index] if index < len(row) else None, key)
        data.append(item)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(str(output_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
